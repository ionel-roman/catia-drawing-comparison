import os
import tkinter as tk
from tkinter import filedialog, messagebox

from extractors.dimension_extractor import extract_dimension_parameters
from core.comparator import compare_by_value_search
from catia.connection import get_catia, open_document


class App:
    def __init__(self, root):
        self.root = root

        self.folder_path = tk.StringVar()
        self.selected_ref = tk.StringVar()
        self.selected_cmp = tk.StringVar()

        # UI
        tk.Button(root, text="Select Folder", command=self.select_folder).pack(pady=5)
        tk.Label(root, textvariable=self.folder_path).pack()

        tk.Label(root, text="Reference Drawing").pack()
        self.ref_menu = tk.OptionMenu(root, self.selected_ref, "")
        self.ref_menu.pack()

        tk.Label(root, text="Compared Drawing").pack()
        self.cmp_menu = tk.OptionMenu(root, self.selected_cmp, "")
        self.cmp_menu.pack()

        tk.Button(root, text="Run Comparison", command=self.run).pack(pady=15)

    def select_folder(self):
        folder = filedialog.askdirectory()
        if not folder:
            return

        self.folder_path.set(folder)

        files = [f for f in os.listdir(folder) if f.lower().endswith(".catdrawing")]

        if not files:
            messagebox.showerror("Error", "No CATDrawing files found")
            return

        self.update_menu(self.ref_menu, self.selected_ref, files)
        self.update_menu(self.cmp_menu, self.selected_cmp, files)

    def update_menu(self, menu, variable, options):
        menu["menu"].delete(0, "end")

        for opt in options:
            menu["menu"].add_command(label=opt, command=lambda v=opt: variable.set(v))

        if options:
            variable.set(options[0])

    def highlight_excel_differences(self, path):
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(path)
        ws = wb.active

        if ws is None:
            return

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        print("Excel headers:", headers)

        try:
            col_val1 = headers.index("value_DRW_1") + 1
            col_val2 = headers.index("value_DRW_2") + 1
            col_status = headers.index("status") + 1
        except ValueError:
            print("Column names not found → skipping highlighting")
            return

        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=col_status).value
            val1 = ws.cell(row=row, column=col_val1).value
            val2 = ws.cell(row=row, column=col_val2).value

            # 🔴 Missing / extra dimensions
            if status != "OK":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red_fill

            # 🟡 Values exist but are different
            elif val1 is not None and val2 is not None:
                try:
                    if abs(float(val1) - float(val2)) > 0:
                        ws.cell(row=row, column=col_val1).fill = yellow_fill
                        ws.cell(row=row, column=col_val2).fill = yellow_fill
                except Exception:
                    pass  # skip if conversion fails

        wb.save(path)
        print("✅ Highlighting applied")
        
    def run(self):
        folder = self.folder_path.get()
        ref = self.selected_ref.get()
        cmp = self.selected_cmp.get()

        if not folder or not ref or not cmp:
            messagebox.showerror("Error", "Selection incomplete")
            return

        path1 = os.path.join(folder, ref)
        path2 = os.path.join(folder, cmp)

        try:
            catia = get_catia()

            doc1 = open_document(catia, path1)
            doc2 = open_document(catia, path2)

            df1 = extract_dimension_parameters(doc1)
            df2 = extract_dimension_parameters(doc2)

            df_result = compare_by_value_search(df1, df2)

            current_dir = os.getcwd()
            reports_dir = os.path.join(current_dir, "reports")
            os.makedirs(reports_dir, exist_ok=True)

            save_path = os.path.join(reports_dir, "output.xlsx")

            # ✅ Save Excel
            df_result.to_excel(save_path, index=False)

            # ✅ Highlight differences
            self.highlight_excel_differences(save_path)

            messagebox.showinfo("Done", f"Saved to:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Error", str(e))